"""Geracao de relatorios: xlsx profissional e terminal formatado."""

import logging
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rich.console import Console
from rich.table import Table

import config
from metricas import METRICAS, INSTRUCOES_DASHBOARD

logger = logging.getLogger(__name__)
console = Console()

BORDA_FINA = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
FILL_HEADER = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FONT_CORPO = Font(size=10, name="Calibri")
FONT_MONO = Font(size=9, name="Consolas")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

FILL_OK = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FILL_ZERO = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
FILL_AVISO = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
FILL_PREENCHER = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_ZEBRA = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
FONT_PREENCHER = Font(italic=True, color="996600", size=10, name="Calibri")

CORES_STATUS: dict[str, str] = {
    "OK": "green",
    "ZERO_LEGITIMO": "yellow",
    "ZERO_SUSPEITO": "red",
    "AUSENTE": "red",
    "NULL": "yellow",
    "TABELA_INEXISTENTE": "dim",
    "ERRO_PERMISSAO": "dim",
    "ERRO": "red",
}


def imprimir_linha(resultado: dict) -> None:
    """Imprime uma linha de resultado no terminal."""
    nome = resultado["nome_metrica"].ljust(35)
    status = resultado["status"]
    valor = resultado["valor_metrica"]

    if valor is not None:
        try:
            valor_fmt = f"{float(valor):,.2f}"
        except (TypeError, ValueError):
            valor_fmt = str(valor)
    else:
        valor_fmt = "NULL"

    cor = CORES_STATUS.get(status, "white")
    console.print(f"  {nome} [{cor}]{valor_fmt:>15}  {status}[/{cor}]")


def imprimir_resumo(resultados: list[dict]) -> None:
    """Imprime tabela resumo no terminal."""
    tabela = Table(title="Resumo da Validacao")
    tabela.add_column("Status", style="bold")
    tabela.add_column("Quantidade", justify="right")

    contagem: dict[str, int] = {}
    for res in resultados:
        contagem[res["status"]] = contagem.get(res["status"], 0) + 1

    for status, qtd in sorted(contagem.items()):
        cor = CORES_STATUS.get(status, "white")
        tabela.add_row(f"[{cor}]{status}[/{cor}]", str(qtd))

    console.print(tabela)


def gerar_xlsx(resultados: list[dict], caminho: Optional[Path] = None) -> Path:
    """Gera xlsx profissional com abas Guia, Input e Output."""
    caminho = caminho or config.ARQUIVO_XLSX
    wb = Workbook()

    _criar_aba_guia(wb)
    _criar_aba_input(wb)
    _criar_aba_output(wb, resultados)

    wb.save(str(caminho))
    console.print(f"\nXLSX salvo em: [bold]{caminho}[/bold]")
    return caminho


def _criar_aba_guia(wb: Workbook) -> None:
    """Cria aba Guia com instrucoes para o usuario."""
    ws = wb.active
    ws.title = "Guia"
    ws.sheet_properties.tabColor = "3498DB"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    titulo_font = Font(bold=True, size=14, name="Calibri", color="1B3A5C")
    subtitulo_font = Font(bold=True, size=11, name="Calibri", color="2C3E50")
    corpo_font = Font(size=10, name="Calibri")

    linhas = [
        ("Validador de Cidades", "Diagnostico automatico do Painel Estrategico Territorial", titulo_font),
        ("", "", None),
        ("Como funciona", "O script le as metricas da aba Input, executa no BigQuery e preenche a aba Output.", subtitulo_font),
        ("", "", None),
        ("Aba Input", "Definicao das metricas a validar. Cada linha = 1 metrica.", subtitulo_font),
        ("pagina_dashboard", "Pagina do dashboard (ex: '1 - Educacao Basica I')", corpo_font),
        ("secao_dashboard", "Secao dentro da pagina (ex: 'CNCA', 'FUNDEB')", corpo_font),
        ("query", "Nome da tabela dbt (ex: 'painel_escola')", corpo_font),
        ("localizacao_gcp", "Caminho completo no BigQuery (projeto.dataset.tabela)", corpo_font),
        ("status_modelo", "Dropdown: ativo / inativo / movido / externo", corpo_font),
        ("coluna_municipio", "Coluna usada para filtrar municipio (ex: 'municipio', 'nome_municipio')", corpo_font),
        ("formato_municipio", "Dropdown: 'cidade - UF' (Andradina - SP), 'cidade' (Andradina), 'CIDADE' (ANDRADINA)", corpo_font),
        ("coluna_estado", "Coluna de estado/UF para filtro adicional (ex: 'sigla_uf', 'uf')", corpo_font),
        ("filtro_por_id", "Dropdown: True/False. True filtra por id IBGE do municipio (evita triplicacao)", corpo_font),
        ("filtro_extra", "Condicao WHERE adicional (ex: \"status = 'Realizado'\")", corpo_font),
        ("coluna_ano", "Coluna de ano para filtro temporal (ex: 'ano', 'ano_tratado'). Vazio = sem filtro de ano", corpo_font),
        ("expressao_metrica", "Expressao SQL ou nome de coluna (ex: 'SUM(quantidade_escola)' ou 'COUNT(DISTINCT id_escola)')", corpo_font),
        ("nome_metrica", "Nome legivel da metrica (ex: 'Escolas', 'FUNDEB Repasse')", corpo_font),
        ("", "", None),
        ("Aba Output", "Resultado automatico. Nao editar colunas azuis.", subtitulo_font),
        ("valor_dashboard (amarelo)", "VOCE PREENCHE: valor visivel no dashboard para comparar", corpo_font),
        ("bate_com_dashboard (amarelo)", "VOCE PREENCHE: SIM / NAO / PARCIAL", corpo_font),
        ("o_que_preciso_do_dashboard", "Instrucao de onde encontrar o valor no dashboard", corpo_font),
        ("", "", None),
        ("Status possiveis", "", subtitulo_font),
        ("OK", "Dados presentes e com valor > 0", corpo_font),
        ("ZERO_LEGITIMO", "Zero confirmado na fonte (municipio nao participa)", corpo_font),
        ("ZERO_SUSPEITO", "Dado pode ter se perdido no pipeline", corpo_font),
        ("AUSENTE", "Municipio nao encontrado na tabela", corpo_font),
        ("", "", None),
        ("Colunas upstream", "Preenchidas com flag --upstream", subtitulo_font),
        ("fonte_upstream", "Tabelas fonte consultadas e contagens", corpo_font),
        ("diagnostico_upstream", "ZERO_CONFIRMADO_FONTE / PERDEU_NO_PAINEL / VALOR_ZERADO", corpo_font),
        ("camada_falha", "Onde o dado se perdeu: nenhuma / painel / politica", corpo_font),
        ("", "", None),
        ("Para adicionar metrica", "Adicione uma nova linha na aba Input preenchendo todos os campos. Na proxima execucao, o script vai inclui-la automaticamente.", subtitulo_font),
        ("Dica", "Use os dropdowns (formato_municipio, filtro_por_id, status_modelo) para evitar erros de digitacao.", corpo_font),
    ]

    for row_idx, (campo, descricao, font) in enumerate(linhas, 1):
        cel_a = ws.cell(row=row_idx, column=1, value=campo)
        cel_b = ws.cell(row=row_idx, column=2, value=descricao)
        if font:
            cel_a.font = font
            cel_b.font = font if font != subtitulo_font else corpo_font
        if font == subtitulo_font:
            cel_a.font = subtitulo_font

    ws.freeze_panes = "A4"


def _estilizar_header(ws, max_col: int) -> None:
    """Aplica estilo ao header de qualquer aba."""
    ws.row_dimensions[1].height = 30
    for col in range(1, max_col + 1):
        celula = ws.cell(row=1, column=col)
        celula.fill = FILL_HEADER
        celula.font = FONT_HEADER
        celula.alignment = ALIGN_CENTER
        celula.border = BORDA_FINA


def _aplicar_bordas(ws, max_row: int, max_col: int) -> None:
    """Aplica bordas a todas as celulas de dados."""
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            celula = ws.cell(row=row, column=col)
            celula.border = BORDA_FINA
            if not celula.font or celula.font == Font():
                celula.font = FONT_CORPO


def _criar_aba_input(wb: Workbook) -> None:
    """Cria aba Input profissional com validacoes e dropdowns."""
    ws = wb.create_sheet("Input")

    colunas = [
        ("pagina_dashboard", 22),
        ("secao_dashboard", 20),
        ("query", 32),
        ("localizacao_gcp", 55),
        ("status_modelo", 14),
        ("coluna_municipio", 18),
        ("formato_municipio", 16),
        ("coluna_estado", 14),
        ("filtro_por_id", 12),
        ("filtro_extra", 28),
        ("coluna_ano", 12),
        ("expressao_metrica", 55),
        ("nome_metrica", 38),
    ]

    for col_idx, (titulo, largura) in enumerate(colunas, 1):
        ws.cell(row=1, column=col_idx, value=titulo)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    _estilizar_header(ws, len(colunas))

    dv_formato = DataValidation(
        type="list", formula1='"cidade - UF,cidade,CIDADE"', allow_blank=True,
    )
    dv_formato.promptTitle = "Formato"
    dv_formato.prompt = "Como o municipio aparece na tabela"
    ws.add_data_validation(dv_formato)

    dv_filtro_id = DataValidation(
        type="list", formula1='"True,False"', allow_blank=True,
    )
    dv_filtro_id.promptTitle = "Filtro por ID"
    dv_filtro_id.prompt = "True evita triplicacao pelo filtro_territorio"
    ws.add_data_validation(dv_filtro_id)

    dv_status = DataValidation(
        type="list", formula1='"ativo,inativo,movido,externo"', allow_blank=True,
    )
    ws.add_data_validation(dv_status)

    for row_idx, metrica in enumerate(METRICAS, 2):
        ws.cell(row=row_idx, column=1, value=metrica["pagina"]).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=2, value=metrica["secao"]).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=3, value=metrica["query"]).font = FONT_MONO
        ws.cell(row=row_idx, column=4, value=metrica["gcp"]).font = FONT_MONO
        cel_status = ws.cell(row=row_idx, column=5, value=metrica["status_modelo"])
        cel_status.alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=6, value=metrica["coluna_municipio"]).font = FONT_MONO
        ws.cell(row=row_idx, column=7, value=metrica.get("formato_municipio", ""))
        ws.cell(row=row_idx, column=8, value=metrica.get("coluna_estado", "")).font = FONT_MONO
        ws.cell(row=row_idx, column=9, value=str(metrica.get("filtro_por_id", ""))).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=10, value=metrica.get("filtro_extra", "")).font = FONT_MONO
        ws.cell(row=row_idx, column=11, value=str(metrica.get("coluna_ano", ""))).font = FONT_MONO
        ws.cell(row=row_idx, column=12, value=metrica["expressao"]).font = FONT_MONO
        ws.cell(row=row_idx, column=13, value=metrica["nome_metrica"]).font = Font(bold=True, size=10, name="Calibri")

        dv_formato.add(ws.cell(row=row_idx, column=7))
        dv_filtro_id.add(ws.cell(row=row_idx, column=9))
        dv_status.add(cel_status)

    _aplicar_bordas(ws, len(METRICAS) + 1, len(colunas))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(METRICAS) + 1}"
    ws.sheet_properties.tabColor = "1B3A5C"


def _criar_aba_output(wb: Workbook, resultados: list[dict]) -> None:
    """Cria aba Output profissional com formatacao condicional."""
    ws = wb.create_sheet("Output")

    colunas = [
        ("municipio", 22),
        ("pagina_dashboard", 22),
        ("secao_dashboard", 20),
        ("query", 30),
        ("nome_metrica", 35),
        ("expressao_metrica", 45),
        ("ano_filtro", 10),
        ("total_registros", 14),
        ("valor_metrica", 20),
        ("valor_dashboard", 18),
        ("bate_com_dashboard", 16),
        ("status", 18),
        ("diagnostico", 55),
        ("o_que_preciso_do_dashboard", 55),
        ("fonte_upstream", 40),
        ("registros_upstream", 16),
        ("diagnostico_upstream", 50),
        ("camada_falha", 22),
        ("query_executada", 65),
        ("data_execucao", 18),
    ]

    for col_idx, (titulo, largura) in enumerate(colunas, 1):
        ws.cell(row=1, column=col_idx, value=titulo)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    _estilizar_header(ws, len(colunas))

    dv_bate = DataValidation(
        type="list", formula1='"SIM,NAO,PARCIAL"', allow_blank=True,
    )
    dv_bate.promptTitle = "Confere?"
    dv_bate.prompt = "O valor do script bate com o dashboard?"
    ws.add_data_validation(dv_bate)

    for row_idx, res in enumerate(resultados, 2):
        ws.cell(row=row_idx, column=1, value=res["municipio"])
        ws.cell(row=row_idx, column=2, value=res["pagina"])
        ws.cell(row=row_idx, column=3, value=res["secao"])
        ws.cell(row=row_idx, column=4, value=res["query"]).font = FONT_MONO
        ws.cell(row=row_idx, column=5, value=res["nome_metrica"]).font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=row_idx, column=6, value=res["expressao"]).font = FONT_MONO
        ws.cell(row=row_idx, column=7, value=res.get("ano_filtro")).alignment = ALIGN_CENTER

        cel_reg = ws.cell(row=row_idx, column=8, value=res["total_registros"])
        cel_reg.alignment = ALIGN_RIGHT
        cel_reg.number_format = "#,##0"

        valor = res["valor_metrica"]
        if valor is not None:
            try:
                valor = float(valor)
            except (TypeError, ValueError):
                pass
        cel_val = ws.cell(row=row_idx, column=9, value=valor)
        cel_val.alignment = ALIGN_RIGHT
        if isinstance(valor, float):
            cel_val.number_format = "#,##0.00"

        cel_dash = ws.cell(row=row_idx, column=10, value=None)
        cel_dash.fill = FILL_PREENCHER
        cel_dash.font = FONT_PREENCHER

        cel_bate = ws.cell(row=row_idx, column=11, value=None)
        cel_bate.fill = FILL_PREENCHER
        cel_bate.alignment = ALIGN_CENTER
        dv_bate.add(cel_bate)

        cel_status = ws.cell(row=row_idx, column=12, value=res["status"])
        cel_status.alignment = ALIGN_CENTER
        if res["status"] == "OK":
            cel_status.fill = FILL_OK
        elif res["status"] in ("ZERO_SUSPEITO", "AUSENTE", "ERRO"):
            cel_status.fill = FILL_ZERO
        elif res["status"] in ("ZERO_LEGITIMO", "NULL"):
            cel_status.fill = FILL_AVISO

        ws.cell(row=row_idx, column=13, value=res["diagnostico"]).alignment = ALIGN_LEFT

        instrucao = INSTRUCOES_DASHBOARD.get(res["nome_metrica"], "")
        cel_inst = ws.cell(row=row_idx, column=14, value=instrucao)
        if instrucao:
            cel_inst.fill = FILL_PREENCHER
            cel_inst.font = FONT_PREENCHER

        ws.cell(row=row_idx, column=15, value=res.get("fonte_upstream", "")).font = FONT_MONO
        cel_up = ws.cell(row=row_idx, column=16, value=res.get("registros_upstream"))
        cel_up.alignment = ALIGN_RIGHT
        if cel_up.value is not None:
            cel_up.number_format = "#,##0"
        ws.cell(row=row_idx, column=17, value=res.get("diagnostico_upstream", "")).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=18, value=res.get("camada_falha", "")).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=19, value=res["query_executada"]).font = FONT_MONO
        ws.cell(row=row_idx, column=20, value=res["data_execucao"]).alignment = ALIGN_CENTER

    max_row = len(resultados) + 1
    _aplicar_bordas(ws, max_row, len(colunas))

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{max_row}"
    ws.sheet_properties.tabColor = "27AE60" if resultados else "E74C3C"


# "Medir o que e mensuravel e tornar mensuravel o que nao e." - Galileu Galilei
