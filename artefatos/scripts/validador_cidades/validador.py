"""Core de validacao: construcao de queries, execucao e classificacao de resultados."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from rich.console import Console

import config
from executor_bq import ExecutorBQ
from metricas import METRICAS as METRICAS_PADRAO
from relatorio import imprimir_linha

logger = logging.getLogger(__name__)
console = Console()


def carregar_metricas_xlsx(caminho: Optional[Path] = None) -> list[dict]:
    """Carrega metricas da aba Input do xlsx, se existir.

    Permite ao usuario adicionar novas metricas direto no xlsx
    sem precisar editar o codigo Python.
    """
    caminho = caminho or config.ARQUIVO_XLSX
    if not caminho.exists():
        return []

    try:
        wb = load_workbook(str(caminho), read_only=True, data_only=True)
    except Exception:
        logger.warning("Falha ao ler xlsx: %s", caminho)
        return []

    if "Input" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Input"]
    headers = [
        str(ws.cell(row=1, column=col).value or "").strip().lower()
        for col in range(1, ws.max_column + 1)
    ]

    mapa_colunas = {
        "pagina_dashboard": "pagina",
        "secao_dashboard": "secao",
        "query": "query",
        "localizacao_gcp": "gcp",
        "status_modelo": "status_modelo",
        "coluna_municipio": "coluna_municipio",
        "formato_municipio": "formato_municipio",
        "coluna_ano": "coluna_ano",
        "expressao_metrica": "expressao",
        "nome_metrica": "nome_metrica",
        "coluna_estado": "coluna_estado",
        "filtro_por_id": "filtro_por_id",
        "filtro_extra": "filtro_extra",
    }

    metricas = []
    for row in range(2, ws.max_row + 1):
        valores = {}
        for col_idx, header in enumerate(headers):
            celula = ws.cell(row=row, column=col_idx + 1).value
            chave = mapa_colunas.get(header)
            if chave and celula is not None:
                valores[chave] = str(celula).strip()

        if not valores.get("query") or not valores.get("expressao"):
            continue

        if valores.get("coluna_ano") in ("", "None", "none"):
            valores["coluna_ano"] = None

        if valores.get("filtro_por_id") in ("True", "true", "1", "sim"):
            valores["filtro_por_id"] = True
        else:
            valores.pop("filtro_por_id", None)

        metricas.append(valores)

    wb.close()

    if metricas:
        logger.info("Carregadas %d metricas do xlsx", len(metricas))

    return metricas


def obter_metricas(xlsx_path: Optional[Path] = None) -> list[dict]:
    """Retorna metricas do xlsx se existir, senao usa as padrao do Python."""
    metricas_xlsx = carregar_metricas_xlsx(xlsx_path)
    if metricas_xlsx:
        console.print(f"  [dim]Metricas carregadas do xlsx ({len(metricas_xlsx)})[/dim]")
        return metricas_xlsx

    console.print(f"  [dim]Metricas padrao do Python ({len(METRICAS_PADRAO)})[/dim]")
    return METRICAS_PADRAO


def parsear_municipio(municipio: str) -> tuple[str, str]:
    """Extrai nome da cidade e UF de 'Andradina - SP'."""
    partes = municipio.rsplit(" - ", 1)
    if len(partes) == 2:
        return partes[0].strip(), partes[1].strip()
    return municipio.strip(), ""


def resolver_id_municipio(executor: ExecutorBQ, municipio: str) -> str:
    """Busca o id_municipio IBGE no filtro_territorio."""
    tabela_ref = f"{config.PROJETO_BQ}.projeto_painel_ministro.painel_escola"
    sql = (
        "SELECT CAST(id AS STRING) as id_mun "
        f"FROM `{tabela_ref}` "
        f"WHERE municipio = '{municipio}' LIMIT 1"
    )
    try:
        rows = executor.executar(sql)
        if rows:
            return rows[0]["id_mun"]
    except Exception:
        logger.warning("Falha ao resolver id_municipio para %s", municipio)
    return ""


def construir_where(
    metrica: dict,
    municipio: str,
    ano: Optional[int] = None,
    id_municipio: str = "",
) -> str:
    """Constroi clausula WHERE para filtrar por municipio."""
    cidade, uf = parsear_municipio(municipio)
    coluna = metrica["coluna_municipio"]
    formato = metrica.get("formato_municipio", "cidade - UF")
    coluna_estado = metrica.get("coluna_estado")
    usar_id = metrica.get("filtro_por_id", False)

    condicoes = []

    if usar_id and id_municipio:
        condicoes.append(f"CAST(id AS STRING) = '{id_municipio}'")
    elif formato == "cidade - UF":
        condicoes.append(f"{coluna} = '{municipio}'")
    elif formato == "CIDADE":
        condicoes.append(f"UPPER({coluna}) = '{cidade.upper()}'")
        if uf and coluna_estado:
            condicoes.append(f"{coluna_estado} = '{uf}'")
    elif formato == "cidade":
        condicoes.append(f"LOWER({coluna}) = '{cidade.lower()}'")
        if uf and coluna_estado:
            condicoes.append(f"{coluna_estado} = '{uf}'")

    filtro_extra = metrica.get("filtro_extra")
    if filtro_extra:
        condicoes.append(filtro_extra)

    col_ano = metrica.get("coluna_ano")
    if col_ano and ano:
        condicoes.append(f"CAST({col_ano} AS INT64) = {ano}")

    return " AND ".join(condicoes)


def construir_query(
    metrica: dict,
    municipio: str,
    ano: Optional[int] = None,
    id_municipio: str = "",
) -> str:
    """Constroi query completa para uma metrica."""
    tabela = metrica["gcp"]
    expressao = metrica["expressao"]
    where = construir_where(metrica, municipio, ano, id_municipio)

    return (
        f"SELECT\n"
        f"  COUNT(*) AS total_registros,\n"
        f"  {expressao} AS valor_metrica\n"
        f"FROM `{tabela}`\n"
        f"WHERE {where}"
    )


def classificar_resultado(
    total_registros: int,
    valor_metrica,
    metrica: dict,
    erro: Optional[str] = None,
) -> tuple[str, str]:
    """Classifica resultado e gera diagnostico."""
    if erro:
        if "Not found" in erro:
            return "TABELA_INEXISTENTE", f"Tabela nao existe no BigQuery: {erro}"
        if "Access Denied" in erro or "permission" in erro.lower():
            return "ERRO_PERMISSAO", f"Sem permissao para acessar: {erro}"
        return "ERRO", f"Erro na execucao: {erro}"

    if total_registros == 0:
        return "AUSENTE", "Municipio nao encontrado nesta tabela."

    if valor_metrica is None:
        return "NULL", "Metrica retornou NULL (campo inexistente ou todos nulos)."

    try:
        valor_num = float(valor_metrica)
    except (TypeError, ValueError):
        return "OK", f"Valor nao numerico: {valor_metrica}"

    if valor_num > 0:
        return "OK", "Dados presentes."

    query_nome = metrica["query"]
    tabelas_cross_join = {
        "painel_pronatec_completo",
        "painel_mulheresmil_completo",
        "painel_sisu",
        "painel_novopac_pacto",
        "painel_novopac_sesu",
    }
    if query_nome in tabelas_cross_join:
        return (
            "ZERO_LEGITIMO",
            "CROSS JOIN gera registros esqueleto com valor 0. "
            "Municipio nao participa deste programa.",
        )

    return "ZERO_SUSPEITO", "Registros existem mas valor e 0. Verificar fonte upstream."


def executar_validacao(
    municipio: str,
    ano: Optional[int] = None,
    dry_run: bool = False,
    upstream: bool = False,
) -> list[dict]:
    """Executa validacao completa para um municipio."""
    from linhagem import buscar_upstream, classificar_upstream

    executor = ExecutorBQ()
    resultados = []
    ano = ano or config.ANO_PADRAO

    console.print(f"\nValidando: [bold]{municipio}[/bold] (ano: {ano})")

    id_municipio = resolver_id_municipio(executor, municipio)
    if id_municipio:
        console.print(f"  id_municipio IBGE: {id_municipio}")
    else:
        console.print("  [yellow]id_municipio nao encontrado[/yellow]")

    if upstream:
        console.print("  [cyan]Modo upstream ativado[/cyan]")

    metricas = obter_metricas()
    console.print()

    for metrica in metricas:
        sql = construir_query(metrica, municipio, ano, id_municipio)

        if dry_run:
            try:
                bytes_est = executor.dry_run(sql)
                console.print(f"  {metrica['nome_metrica']}: ~{bytes_est / 1024:.0f} KB")
            except Exception as exc:
                console.print(f"  {metrica['nome_metrica']}: ERRO dry-run ({exc})")
            continue

        resultado = {
            "municipio": municipio,
            "pagina": metrica["pagina"],
            "secao": metrica["secao"],
            "query": metrica["query"],
            "gcp": metrica["gcp"],
            "nome_metrica": metrica["nome_metrica"],
            "expressao": metrica["expressao"],
            "ano_filtro": ano if metrica.get("coluna_ano") else None,
            "total_registros": 0,
            "valor_metrica": None,
            "status": "",
            "diagnostico": "",
            "query_executada": sql,
            "data_execucao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fonte_upstream": "",
            "registros_upstream": None,
            "diagnostico_upstream": "",
            "camada_falha": "",
        }

        try:
            rows = executor.executar(sql)
            if rows:
                resultado["total_registros"] = rows[0].get("total_registros", 0)
                resultado["valor_metrica"] = rows[0].get("valor_metrica")

            status, diag = classificar_resultado(
                resultado["total_registros"],
                resultado["valor_metrica"],
                metrica,
            )
            resultado["status"] = status
            resultado["diagnostico"] = diag

        except Exception as exc:
            status, diag = classificar_resultado(0, None, metrica, str(exc))
            resultado["status"] = status
            resultado["diagnostico"] = diag

        imprimir_linha(resultado)

        if upstream and resultado["status"] in (
            "ZERO_SUSPEITO", "ZERO_LEGITIMO", "AUSENTE", "NULL"
        ):
            _executar_upstream(
                executor, metrica, resultado, id_municipio, municipio, ano
            )

        resultados.append(resultado)

    if not dry_run:
        console.print(f"\n[dim]{executor.resumo_custo()}[/dim]")

    return resultados


def _executar_upstream(
    executor: ExecutorBQ,
    metrica: dict,
    resultado: dict,
    id_municipio: str,
    municipio: str,
    ano: Optional[int],
) -> None:
    """Executa verificacao upstream para uma metrica zerada."""
    from linhagem import buscar_upstream, classificar_upstream

    upstream_res = buscar_upstream(
        executor,
        metrica["query"],
        id_municipio,
        municipio,
        ano if metrica.get("coluna_ano") else None,
    )

    fontes_str = []
    total_up = 0
    for r in upstream_res:
        reg = r.get("registros")
        nome = r.get("fonte", "?")
        if reg is not None:
            fontes_str.append(f"{nome}={reg}")
            total_up += reg
        elif r.get("erro"):
            fontes_str.append(f"{nome}=ERRO")
        else:
            fontes_str.append(f"{nome}=?")

    resultado["fonte_upstream"] = " | ".join(fontes_str)
    resultado["registros_upstream"] = total_up

    up_status, up_diag = classificar_upstream(
        resultado["total_registros"],
        resultado["valor_metrica"],
        upstream_res,
    )
    resultado["diagnostico_upstream"] = up_diag

    if up_status == "ZERO_CONFIRMADO_FONTE":
        resultado["camada_falha"] = "nenhuma (zero real)"
        resultado["status"] = "ZERO_LEGITIMO"
        resultado["diagnostico"] = up_diag
    elif up_status == "PERDEU_NO_PAINEL":
        resultado["camada_falha"] = "painel (JOIN/filtro)"
        resultado["status"] = "ZERO_SUSPEITO"
        resultado["diagnostico"] = up_diag
    elif up_status == "VALOR_ZERADO_NO_PAINEL":
        resultado["camada_falha"] = "painel (expressao/filtro)"
    else:
        resultado["camada_falha"] = up_status.lower()

    cor_up = "green" if total_up == 0 else "red"
    console.print(
        f"    [dim]upstream:[/dim] [{cor_up}]{resultado['fonte_upstream']}[/{cor_up}]"
        f" -> [bold]{up_status}[/bold]"
    )


# "Aquele que tem um porque para viver pode suportar quase qualquer como." - Friedrich Nietzsche
